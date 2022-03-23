"""
数据读写操作
"""

from openpyxl import load_workbook


def find_none(l):
    """
    返回列表中的非None项
    :param l: 列表
    :return:new_l: 列表
    """

    new_l = []

    for _ in l:  # 遍历列表
        if _ is not None:  # 如果不是 none
            new_l.append(_)  # 就加入新列表

    return new_l


def case_by_excel(path):
    """
    读取execl文件内容，并返回测试用例
    :param path: excel 文件的路径
    :return:
    """

    wb = load_workbook(path)  # 整个工作簿（整个文件）

    for ws in wb.worksheets:  # 1. 遍历sheet
        # 遍历sheet，创建新的套件
        _suite = {  # 套件的格式
            "info": {  # 套件的信息
                "name": ws.title,
            },
            "cases": {},  # 套件中的测试用例
        }
        _cases = _suite["cases"]

        new_case_id = 0  # 给用例设置ID

        for (step_id, step_name, keyword, *args,) in ws.iter_rows(
            min_row=2, values_only=True
        ):  # 2. 遍历行
            # 3.解析内容
            if step_id < 0:  # 用例info部分
                new_case_id += 1  # 新的测试用例
                # 根据id获取用例，没有就创建一个新的
                new_case = _cases.get(new_case_id, {"info": {}, "steps": []})
                new_case["info"][keyword] = args[0]

            else:  # 用例 step部分
                new_case = _cases.get(new_case_id, {"info": {}, "steps": []})
                # 如果参数是None就忽略
                new_case["steps"].append((keyword, [_ for _ in args if _]))

            _cases[new_case_id] = new_case  # 把处理好的信息放回套件中
        yield _suite
